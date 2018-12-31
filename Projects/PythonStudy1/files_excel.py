import openpyxl
import os, sys
#########################################################
#Set script path and remove trailing \ if path is c:\
#########################################################
scriptPath = os.path.dirname(sys.argv[0])
if scriptPath[-1:]=='\\':
    scriptPath = scriptPath.strip('\\')
#########################################################
'''
import requests
#get file and save it in variable --> res
res = requests.get('http://autbor.com/example.xlsx')
res.raise_for_status()

#DOWNLOAD FILE!
playfile = open('%s\\example.xlsx' % (scriptPath), 'wb') # open file in wb --- write-binary mode
for chunk in res.iter_content(100000):  # specify number of bytes per chunk
    playfile.write(chunk)

playfile.close()
'''
'''
workbook
sheet
columns
row
cells

pip install openpyxl
'''
os.chdir(scriptPath)
print(os.getcwd())
workbook = openpyxl.load_workbook('example.xlsx')
print(type(workbook))

print(workbook.get_sheet_names())
sheet = workbook.get_sheet_by_name('Sheet1')
print(sheet['A1'].value) # will return the actual data type in excel but you can pass it to str()
print(sheet.cell(row=1, column=2).value)

for i in range(1,8):
        print(i, sheet.cell(row=i, column=2).value)