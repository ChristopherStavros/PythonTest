import openpyxl
import os, sys
#########################################################
#Set script path and remove trailing \ if path is c:\
#########################################################
scriptPath = os.path.dirname(sys.argv[0])
if scriptPath[-1:]=='\\':
    scriptPath = scriptPath.strip('\\')
#########################################################
os.chdir(scriptPath)
print(os.getcwd())
#########################################################

wb = openpyxl.Workbook()
print(wb.sheetnames)
sheet = wb.get_sheet_by_name('Sheet')
print(sheet['A1'].value)
sheet['A1'].value = 42
sheet['A2'].value = 'Hello'
print(sheet['A1'].value)

wb.save('example2.xlsx')
sheet2 = wb.create_sheet()
print(wb.sheetnames)
sheet2.title = 'My new sheet name'
print(wb.sheetnames)
wb.save('example3.xlsx')

wb.create_sheet(index=0, title = "my Oher Sheet") #index 0 make this the first sheet
print(wb.sheetnames)
wb.save('example3.xlsx')